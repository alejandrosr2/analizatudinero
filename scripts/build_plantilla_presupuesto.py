# Genera la "Plantilla de presupuesto mensual" de Analiza Tu Dinero.
# Salida: scripts/output/plantilla-presupuesto-mensual-analizatudinero.xlsx
# (fuera de public/: ya no se distribuye gratis en la web; sirve de base
#  para módulos del producto premium "Controla tu dinero en 30 días").
# Regenerar: python scripts/build_plantilla_presupuesto.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference

NAVY = "0F2E4C"
NAVY_100 = "DCE8F3"
MINT_100 = "DDF3E8"
MINT_600 = "2F7E5E"
SLATE_BORDER = "C7D2DC"
SLATE_TEXT = "475569"
AMBER_BG = "FEF3C7"
AMBER_TEXT = "92400E"

EURO = '#,##0.00 "€"'
EURO_CALC = '#,##0.00 "€";[Red]-#,##0.00 "€";"—"'
PCT = "0.0%"

F_BASE = Font(name="Arial", size=10)
F_BOLD = Font(name="Arial", size=10, bold=True)
F_TITLE = Font(name="Arial", size=15, bold=True, color="FFFFFF")
F_H2 = Font(name="Arial", size=11, bold=True, color=NAVY)
F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_NOTE = Font(name="Arial", size=9, italic=True, color=SLATE_TEXT)

FILL_NAVY = PatternFill("solid", start_color=NAVY)
FILL_NAVY100 = PatternFill("solid", start_color=NAVY_100)
FILL_INPUT = PatternFill("solid", start_color=MINT_100)
FILL_AMBER = PatternFill("solid", start_color=AMBER_BG)

THIN = Side(style="thin", color=SLATE_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_LINE = Border(top=Side(style="medium", color=NAVY))

CATEGORIAS = [
    "Vivienda", "Suministros", "Alimentación", "Transporte",
    "Salud y cuidado personal", "Ocio y restaurantes", "Suscripciones",
    "Deudas", "Ropa y hogar", "Otros gastos",
]
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

REG_FIRST, REG_LAST = 5, 304  # filas de datos del registro de gastos


def style(cell, font=F_BASE, fill=None, fmt=None, align=None, border=None):
    cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def title_bar(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=text)
    style(c, F_TITLE, FILL_NAVY, align=Alignment(vertical="center", indent=1))
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = FILL_NAVY
    ws.row_dimensions[1].height = 30


def input_cell(ws, ref, value=None, fmt=EURO):
    c = ws[ref]
    if value is not None:
        c.value = value
    style(c, F_BASE, FILL_INPUT, fmt, border=BORDER)
    return c


wb = Workbook()

# ---------------------------------------------------------------- Empieza aquí
ws = wb.active
ws.title = "Empieza aquí"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 4
ws.column_dimensions["C"].width = 92
ws.column_dimensions["D"].width = 3
title_bar(ws, "Analiza Tu Dinero — Plantilla de presupuesto mensual (versión gratuita)", 4)

rows = [
    ("h", "Cómo usar esta plantilla"),
    ("s", "1.", "En la hoja «Configuración», escribe el mes, revisa las categorías de gasto (puedes renombrarlas) y fija tu objetivo de ahorro."),
    ("s", "2.", "En la hoja «Presupuesto», rellena tus ingresos y lo que PREVÉS gastar en cada categoría (celdas verdes)."),
    ("s", "3.", "Durante el mes, apunta cada gasto en la hoja «Registro de gastos». La columna REAL del presupuesto se actualiza sola."),
    ("s", "4.", "Al cerrar el mes, copia tus totales en la hoja «Resumen anual» para ver la evolución de tu ahorro."),
    ("h", "Leyenda de colores"),
    ("leg_input", "", "Celdas verdes: escribe aquí tus datos."),
    ("leg_calc", "", "Celdas blancas con cifras: se calculan solas, no las edites."),
    ("h", "Tres consejos para que dure más de dos semanas"),
    ("s", "•", "No busques precisión de céntimo: con apuntar los gastos una o dos veces por semana es suficiente."),
    ("s", "•", "Si una categoría se desvía todos los meses, cambia el presupuesto de esa categoría, no te castigues."),
    ("s", "•", "Revisa el resumen 10 minutos a final de mes: es la parte que de verdad cambia las cosas."),
    ("aviso", "", "Contenido educativo, no asesoramiento financiero. Esta plantilla es una herramienta de organización personal: "
                  "no tiene en cuenta tu situación completa, no garantiza ningún resultado económico y sus cifras son las que tú introduces."),
    ("note", "", "Más calculadoras, guías y plantillas en analizatudinero.com · © Analiza Tu Dinero"),
]
r = 3
for row in rows:
    kind = row[0]
    if kind == "h":
        style(ws.cell(row=r, column=2, value=row[1]), F_H2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1
    elif kind == "s":
        style(ws.cell(row=r, column=2, value=row[1]), F_BOLD, align=Alignment(vertical="top"))
        style(ws.cell(row=r, column=3, value=row[2]), F_BASE, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 26
        r += 1
    elif kind == "leg_input":
        style(ws.cell(row=r, column=2), fill=FILL_INPUT, border=BORDER)
        style(ws.cell(row=r, column=3, value=row[2]), F_BASE)
        r += 1
    elif kind == "leg_calc":
        style(ws.cell(row=r, column=2), border=BORDER)
        style(ws.cell(row=r, column=3, value=row[2]), F_BASE)
        r += 2
    elif kind == "aviso":
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=3)
        c = ws.cell(row=r, column=2, value=row[2])
        style(c, Font(name="Arial", size=9, color=AMBER_TEXT), FILL_AMBER,
              align=Alignment(wrap_text=True, vertical="center", indent=1))
        ws.row_dimensions[r].height = 24
        ws.row_dimensions[r + 1].height = 24
        r += 3
    elif kind == "note":
        style(ws.cell(row=r, column=3, value=row[2]), F_NOTE)
        r += 1

# ---------------------------------------------------------------- Configuración
ws = wb.create_sheet("Configuración")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 60
title_bar(ws, "Configuración", 3)

style(ws["A3"], F_H2); ws["A3"] = "Periodo"
style(ws["A4"], F_BASE, border=BORDER); ws["A4"] = "Mes"
input_cell(ws, "B4", "Junio", fmt="@")
style(ws["A5"], F_BASE, border=BORDER); ws["A5"] = "Año"
input_cell(ws, "B5", 2026, fmt="0")

style(ws["A7"], F_H2); ws["A7"] = "Categorías de gasto"
style(ws["C7"], F_NOTE); ws["C7"] = "Puedes renombrarlas: el presupuesto y los desplegables del registro se actualizan solos."
for i, cat in enumerate(CATEGORIAS):
    input_cell(ws, f"A{8 + i}", cat, fmt="@")

style(ws["A20"], F_H2); ws["A20"] = "Objetivo"
style(ws["A21"], F_BASE, border=BORDER); ws["A21"] = "Objetivo de ahorro mensual (€)"
input_cell(ws, "B21", 100)
style(ws["C21"], F_NOTE); ws["C21"] = "Cifra de ejemplo: cámbiala por la tuya. Si no lo tienes claro, la guía de capacidad de ahorro de la web te ayuda."

wb.defined_names.add(DefinedName("Categorias", attr_text="'Configuración'!$A$8:$A$17"))

# ---------------------------------------------------------------- Presupuesto
ws = wb.create_sheet("Presupuesto")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = MINT_600
for col, w in {"A": 28, "B": 15, "C": 15, "D": 15, "E": 14, "F": 2}.items():
    ws.column_dimensions[col].width = w
title_bar(ws, "Presupuesto mensual", 6)
style(ws["A2"], F_NOTE)
ws["A2"] = '=CONCATENATE("Periodo: ", Configuración!B4, " ", Configuración!B5)'

# Ingresos
style(ws["A4"], F_H2); ws["A4"] = "Ingresos"
for ref, txt in (("A5", "Concepto"), ("B5", "Previsto"), ("C5", "Real")):
    ws[ref] = txt
    style(ws[ref], F_HEAD, FILL_NAVY, align=Alignment(horizontal="center"), border=BORDER)
ws["A5"].alignment = Alignment(horizontal="left", indent=1)
for i, concepto in enumerate(["Nómina / ingreso principal", "Ingresos extra", "Otros ingresos"]):
    r = 6 + i
    style(ws.cell(row=r, column=1, value=concepto), F_BASE, border=BORDER)
    input_cell(ws, f"B{r}", 0)
    input_cell(ws, f"C{r}", 0)
style(ws["A9"], F_BOLD, border=BORDER); ws["A9"] = "Total ingresos"
style(ws["B9"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws["B9"] = "=SUM(B6:B8)"
style(ws["C9"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws["C9"] = "=SUM(C6:C8)"

# Gastos
style(ws["A11"], F_H2); ws["A11"] = "Gastos"
heads = [("A12", "Categoría"), ("B12", "Previsto"), ("C12", "Real (automático)"),
         ("D12", "Diferencia"), ("E12", "% de ingresos")]
for ref, txt in heads:
    ws[ref] = txt
    style(ws[ref], F_HEAD, FILL_NAVY, align=Alignment(horizontal="center", wrap_text=True), border=BORDER)
ws["A12"].alignment = Alignment(horizontal="left", indent=1)
GASTO_FIRST, GASTO_LAST = 13, 12 + len(CATEGORIAS)
for i in range(len(CATEGORIAS)):
    r = GASTO_FIRST + i
    style(ws.cell(row=r, column=1, value=f"=Configuración!A{8 + i}"), F_BASE, border=BORDER)
    input_cell(ws, f"B{r}", 0)
    style(ws[f"C{r}"], F_BASE, fmt=EURO_CALC, border=BORDER)
    ws[f"C{r}"] = f"=SUMIFS('Registro de gastos'!$D${REG_FIRST}:$D${REG_LAST},'Registro de gastos'!$B${REG_FIRST}:$B${REG_LAST},$A{r})"
    style(ws[f"D{r}"], F_BASE, fmt=EURO_CALC, border=BORDER); ws[f"D{r}"] = f"=B{r}-C{r}"
    style(ws[f"E{r}"], F_BASE, fmt=PCT, border=BORDER); ws[f"E{r}"] = f"=IF($C$9=0,0,C{r}/$C$9)"
TOT = GASTO_LAST + 1
style(ws[f"A{TOT}"], F_BOLD, border=BORDER); ws[f"A{TOT}"] = "Total gastos"
for col, formula in (("B", f"=SUM(B{GASTO_FIRST}:B{GASTO_LAST})"), ("C", f"=SUM(C{GASTO_FIRST}:C{GASTO_LAST})"),
                     ("D", f"=B{TOT}-C{TOT}")):
    style(ws[f"{col}{TOT}"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws[f"{col}{TOT}"] = formula
style(ws[f"E{TOT}"], F_BOLD, fmt=PCT, border=BORDER); ws[f"E{TOT}"] = f"=IF($C$9=0,0,C{TOT}/$C$9)"

# Resumen del mes
R0 = TOT + 2
style(ws[f"A{R0}"], F_H2); ws[f"A{R0}"] = "Resumen del mes"
resumen = [
    ("Ingresos reales", "=C9", EURO_CALC),
    ("Gastos reales", f"=C{TOT}", EURO_CALC),
    ("Margen del mes (ingresos − gastos)", f"=C9-C{TOT}", EURO_CALC),
    ("% gastado", f"=IF(C9=0,0,C{TOT}/C9)", PCT),
    ("% disponible", f"=IF(C9=0,0,1-C{TOT}/C9)", PCT),
    ("Objetivo de ahorro (Configuración)", "=Configuración!B21", EURO_CALC),
]
for i, (label, formula, fmt) in enumerate(resumen):
    r = R0 + 1 + i
    style(ws.cell(row=r, column=1, value=label), F_BASE, border=BORDER)
    style(ws.cell(row=r, column=2, value=formula), F_BOLD, fmt=fmt, border=BORDER)
MARGEN = f"B{R0 + 3}"
ESTADO = R0 + 7
style(ws.cell(row=ESTADO, column=1, value="Lectura del mes"), F_BASE, border=BORDER)
ws.merge_cells(start_row=ESTADO, start_column=2, end_row=ESTADO, end_column=5)
c = ws.cell(row=ESTADO, column=2)
c.value = (f'=IF(C9=0,"Rellena tus ingresos reales para ver la lectura del mes.",'
           f'IF({MARGEN}<0,"Este mes los gastos superan a los ingresos. Revisa las categorías con mayor desvío.",'
           f'IF({MARGEN}>=Configuración!B21,"Margen por encima de tu objetivo de ahorro. Decide su destino antes de que se gaste solo.",'
           f'"Margen positivo, pero por debajo de tu objetivo de ahorro. Mira la columna Diferencia para localizar desvíos.")))')
style(c, F_BASE, FILL_NAVY100, align=Alignment(wrap_text=True, vertical="center", indent=1))
ws.row_dimensions[ESTADO].height = 30
style(ws.cell(row=ESTADO + 2, column=1,
              value="Las celdas verdes son tuyas; el resto se calcula solo. Estimaciones orientativas: no es asesoramiento financiero."), F_NOTE)

ws.conditional_formatting.add(
    f"D{GASTO_FIRST}:D{TOT}",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(name="Arial", size=10, color="C00000")))
ws.conditional_formatting.add(
    MARGEN, CellIsRule(operator="lessThan", formula=["0"], font=Font(name="Arial", size=10, bold=True, color="C00000")))

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Previsto vs real por categoría"
chart.x_axis.delete = False
chart.y_axis.delete = False
chart.height = 8.5
chart.width = 21
data = Reference(ws, min_col=2, max_col=3, min_row=12, max_row=GASTO_LAST)
cats = Reference(ws, min_col=1, min_row=GASTO_FIRST, max_row=GASTO_LAST)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = NAVY
chart.series[1].graphicalProperties.solidFill = "7FD1AE"
chart.legend.position = "b"
ws.add_chart(chart, f"A{ESTADO + 4}")

# ---------------------------------------------------------------- Registro de gastos
ws = wb.create_sheet("Registro de gastos")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = MINT_600
for col, w in {"A": 14, "B": 26, "C": 38, "D": 14, "E": 2, "F": 18}.items():
    ws.column_dimensions[col].width = w
title_bar(ws, "Registro de gastos", 6)
style(ws["A2"], F_NOTE)
ws["A2"] = "Apunta cada gasto del mes. La columna Categoría tiene desplegable y la hoja Presupuesto se actualiza sola. Borra las filas de ejemplo."
ws["A3"] = ""
for ref, txt in (("A4", "Fecha"), ("B4", "Categoría"), ("C4", "Descripción"), ("D4", "Importe")):
    ws[ref] = txt
    style(ws[ref], F_HEAD, FILL_NAVY, align=Alignment(horizontal="center"), border=BORDER)
ejemplos = [
    ("03/06/2026", "Alimentación", "Compra semanal (ejemplo: bórralo)", 52.30),
    ("05/06/2026", "Transporte", "Abono mensual (ejemplo: bórralo)", 40.00),
    ("08/06/2026", "Ocio y restaurantes", "Cena con amigos (ejemplo: bórralo)", 24.50),
]
for r in range(REG_FIRST, REG_LAST + 1):
    style(ws.cell(row=r, column=1), F_BASE, FILL_INPUT, "DD/MM/YYYY", border=BORDER)
    style(ws.cell(row=r, column=2), F_BASE, FILL_INPUT, "@", border=BORDER)
    style(ws.cell(row=r, column=3), F_BASE, FILL_INPUT, "@", border=BORDER)
    style(ws.cell(row=r, column=4), F_BASE, FILL_INPUT, EURO, border=BORDER)
for i, (fecha, cat, desc, imp) in enumerate(ejemplos):
    r = REG_FIRST + i
    d, m, y = fecha.split("/")
    from datetime import date
    ws.cell(row=r, column=1, value=date(int(y), int(m), int(d)))
    ws.cell(row=r, column=2, value=cat)
    ws.cell(row=r, column=3, value=desc)
    ws.cell(row=r, column=4, value=imp)
style(ws["F4"], F_HEAD, FILL_NAVY, align=Alignment(horizontal="center"), border=BORDER); ws["F4"] = "Total registrado"
style(ws["F5"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws["F5"] = f"=SUM(D{REG_FIRST}:D{REG_LAST})"

dv = DataValidation(type="list", formula1="=Categorias", allow_blank=True,
                    promptTitle="Categoría", prompt="Elige una categoría de la lista (se definen en Configuración).",
                    errorTitle="Categoría no válida", error="Elige una categoría del desplegable o ajusta la lista en Configuración.")
ws.add_data_validation(dv)
dv.add(f"B{REG_FIRST}:B{REG_LAST}")
dv_imp = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True,
                        errorTitle="Importe no válido", error="Escribe un importe en euros igual o mayor que 0.")
ws.add_data_validation(dv_imp)
dv_imp.add(f"D{REG_FIRST}:D{REG_LAST}")
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- Resumen anual
ws = wb.create_sheet("Resumen anual")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY
for col, w in {"A": 14, "B": 15, "C": 15, "D": 15, "E": 18, "F": 12}.items():
    ws.column_dimensions[col].width = w
title_bar(ws, "Resumen anual", 6)
style(ws["A2"], F_NOTE)
ws["A2"] = "Al cerrar cada mes, copia aquí los totales reales de tu hoja Presupuesto (ingresos y gastos)."
heads = [("A4", "Mes"), ("B4", "Ingresos"), ("C4", "Gastos"), ("D4", "Ahorro"),
         ("E4", "Ahorro acumulado"), ("F4", "% ahorro")]
for ref, txt in heads:
    ws[ref] = txt
    style(ws[ref], F_HEAD, FILL_NAVY, align=Alignment(horizontal="center", wrap_text=True), border=BORDER)
for i, mes in enumerate(MESES):
    r = 5 + i
    style(ws.cell(row=r, column=1, value=mes), F_BASE, border=BORDER)
    input_cell(ws, f"B{r}", 0)
    input_cell(ws, f"C{r}", 0)
    style(ws[f"D{r}"], F_BASE, fmt=EURO_CALC, border=BORDER); ws[f"D{r}"] = f"=B{r}-C{r}"
    style(ws[f"E{r}"], F_BASE, fmt=EURO_CALC, border=BORDER)
    ws[f"E{r}"] = "=D5" if r == 5 else f"=E{r - 1}+D{r}"
    style(ws[f"F{r}"], F_BASE, fmt=PCT, border=BORDER); ws[f"F{r}"] = f"=IF(B{r}=0,0,D{r}/B{r})"
style(ws["A17"], F_BOLD, border=BORDER); ws["A17"] = "Total año"
for col in "BCD":
    style(ws[f"{col}17"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws[f"{col}17"] = f"=SUM({col}5:{col}16)"
style(ws["E17"], F_BOLD, fmt=EURO_CALC, border=BORDER); ws["E17"] = "=E16"
style(ws["F17"], F_BOLD, fmt=PCT, border=BORDER); ws["F17"] = "=IF(B17=0,0,D17/B17)"
ws.conditional_formatting.add(
    "D5:E17", CellIsRule(operator="lessThan", formula=["0"], font=Font(name="Arial", size=10, color="C00000")))

lchart = LineChart()
lchart.title = "Ahorro acumulado"
lchart.style = 12
lchart.height = 8
lchart.width = 21
lchart.x_axis.delete = False
lchart.y_axis.delete = False
ldata = Reference(ws, min_col=5, min_row=4, max_row=16)
lcats = Reference(ws, min_col=1, min_row=5, max_row=16)
lchart.add_data(ldata, titles_from_data=True)
lchart.set_categories(lcats)
lchart.series[0].graphicalProperties.line.solidFill = MINT_600
lchart.series[0].graphicalProperties.line.width = 28000
lchart.series[0].smooth = False
lchart.legend = None
ws.add_chart(lchart, "A20")

import os
out = os.path.join(os.path.dirname(__file__), "output",
                   "plantilla-presupuesto-mensual-analizatudinero.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("OK:", os.path.abspath(out))
